"""
Export utilities: generate PDF and Excel financial reports/statements.
"""
import io
import datetime as dt

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
)
from reportlab.lib.styles import getSampleStyleSheet

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def generate_pdf_report(user_name: str, month: str, summary: dict, expenses: list) -> bytes:
    """Build a PDF financial report and return raw bytes."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2 * cm, bottomMargin=2 * cm)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("AI Personal Finance & Expense Tracker", styles["Title"]))
    elements.append(Paragraph(f"Financial Report — {month}", styles["Heading2"]))
    elements.append(Paragraph(f"Prepared for: {user_name}", styles["Normal"]))
    elements.append(Paragraph(f"Generated on: {dt.datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", styles["Normal"]))
    elements.append(Spacer(1, 0.6 * cm))

    summary_data = [
        ["Metric", "Amount (₹)"],
        ["Total Income", f"{summary.get('total_income', 0):,.2f}"],
        ["Total Expenses", f"{summary.get('total_expenses', 0):,.2f}"],
        ["Savings", f"{summary.get('savings', 0):,.2f}"],
        ["Financial Score", f"{summary.get('financial_score', 0)}/100"],
    ]
    summary_table = Table(summary_data, colWidths=[8 * cm, 8 * cm])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ("PADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.8 * cm))

    if summary.get("summary_text"):
        elements.append(Paragraph("AI Summary & Recommendations", styles["Heading2"]))
        elements.append(Paragraph(summary["summary_text"], styles["Normal"]))
        elements.append(Spacer(1, 0.8 * cm))

    elements.append(Paragraph("Recent Transactions", styles["Heading2"]))
    tx_data = [["Date", "Category", "Description", "Amount (₹)"]]
    for e in expenses[:30]:
        tx_data.append([
            e["date"].strftime("%Y-%m-%d") if isinstance(e["date"], dt.datetime) else str(e["date"]),
            e["category"], e.get("description") or "-", f"{e['amount']:,.2f}",
        ])
    tx_table = Table(tx_data, colWidths=[3 * cm, 3.5 * cm, 6 * cm, 3.5 * cm])
    tx_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("PADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(tx_table)

    doc.build(elements)
    return buffer.getvalue()


def generate_excel_report(month: str, summary: dict, expenses: list, incomes: list) -> bytes:
    """Build an Excel workbook with Summary, Expenses, and Income sheets."""
    wb = openpyxl.Workbook()

    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Summary sheet
    ws = wb.active
    ws.title = "Summary"
    ws.append(["AI Personal Finance & Expense Tracker"])
    ws.append([f"Report Month: {month}"])
    ws.append([])
    ws.append(["Metric", "Value"])
    for cell in ws[4]:
        cell.fill = header_fill
        cell.font = header_font
    rows = [
        ("Total Income", summary.get("total_income", 0)),
        ("Total Expenses", summary.get("total_expenses", 0)),
        ("Savings", summary.get("savings", 0)),
        ("Financial Score", summary.get("financial_score", 0)),
    ]
    for r in rows:
        ws.append(r)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18

    # Expenses sheet
    ws2 = wb.create_sheet("Expenses")
    ws2.append(["Date", "Category", "Description", "Payment Mode", "Amount", "Anomaly"])
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
    for e in expenses:
        ws2.append([
            e["date"].strftime("%Y-%m-%d") if isinstance(e["date"], dt.datetime) else str(e["date"]),
            e["category"], e.get("description") or "-", e.get("payment_mode", "-"),
            e["amount"], "Yes" if e.get("is_anomaly") else "No",
        ])
    for col in "ABCDEF":
        ws2.column_dimensions[col].width = 16

    # Income sheet
    ws3 = wb.create_sheet("Income")
    ws3.append(["Date", "Source", "Description", "Amount"])
    for cell in ws3[1]:
        cell.fill = header_fill
        cell.font = header_font
    for i in incomes:
        ws3.append([
            i["date"].strftime("%Y-%m-%d") if isinstance(i["date"], dt.datetime) else str(i["date"]),
            i["source"], i.get("description") or "-", i["amount"],
        ])
    for col in "ABCD":
        ws3.column_dimensions[col].width = 16

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
